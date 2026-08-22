import requests
from bs4 import BeautifulSoup
import csv
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def scrape_and_save():
    url = "http://quotes.toscrape.com"
    response = requests.get(url)
    
    if response.status_code != 200:
        print(f"Failed to fetch page. Status code: {response.status_code}")
        return

    soup = BeautifulSoup(response.text, 'html.parser')
    quotes_elements = soup.find_all('div', class_='quote')
    
    data_list = []
    
    for idx, quote_el in enumerate(quotes_elements, start=1):
        text = quote_el.find('span', class_='text').text.strip("“”")
        author = quote_el.find('small', class_='author').text.strip()
        tags = [tag.text for tag in quote_el.find_all('a', class_='tag')]
        
        data_list.append({
            "ID": idx,
            "Quote": text,
            "Author": author,
            "Tags": ", ".join(tags)
        })

    # 1. Export to CSV
    with open("quotes.csv", mode="w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["ID", "Quote", "Author", "Tags"])
        writer.writeheader()
        writer.writerows(data_list)

    # 2. Export to JSON
    with open("quotes.json", mode="w", encoding="utf-8") as f:
        json.dump(data_list, f, indent=4, ensure_ascii=False)

    # 3. Export to Styled Excel (.xlsx)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotes Data"

    # Add Headers
    headers = ["ID", "Quote", "Author", "Tags"]
    ws.append(headers)

    # Style Header (Dark Blue Background + White Bold Text)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_num in range(1, 5):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 26

    # Add Rows
    for item in data_list:
        ws.append([item["ID"], item["Quote"], item["Author"], item["Tags"]])

    # Style Data Cells & Enable Text Wrapping
    data_font = Font(name="Segoe UI", size=10)
    wrap_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=len(data_list)+1, min_col=1, max_col=4):
        for cell in row:
            cell.font = data_font
            if cell.column == 1:
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align

    # Set Column Widths so text is never hidden
    ws.column_dimensions['A'].width = 8    # ID
    ws.column_dimensions['B'].width = 50   # Quote
    ws.column_dimensions['C'].width = 22   # Author
    ws.column_dimensions['D'].width = 30   # Tags

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    wb.save("quotes.xlsx")
    
    print("Success: quotes.csv, quotes.json, and formatted quotes.xlsx generated!")

if __name__ == "__main__":
    scrape_and_save()