import requests
from bs4 import BeautifulSoup
import csv
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def scrape_books(pages_to_scrape=3):
    base_url = "http://books.toscrape.com/catalogue/page-{}.html"
    data_list = []
    item_id = 1

    # Text rating to number mapping
    rating_map = {
        "One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5
    }

    print(f"Starting scraper for {pages_to_scrape} pages...")

    for page in range(1, pages_to_scrape + 1):
        url = base_url.format(page)
        response = requests.get(url)
        
        if response.status_code != 200:
            print(f"Failed to fetch page {page}")
            break

        soup = BeautifulSoup(response.text, 'html.parser')
        books = soup.find_all('article', class_='product_pod')

        for book in books:
            title = book.h3.a['title']
            price_text = book.find('p', class_='price_color').text.replace('£', '').replace('Â', '').strip()
            
            # Extract Star Rating from CSS class
            rating_class = book.find('p', class_='star-rating')['class'][1]
            rating = rating_map.get(rating_class, 0)

            # Extract Stock Availability
            availability = book.find('p', class_='instock availability').text.strip()
            
            # Full Product Link
            link = "http://books.toscrape.com/catalogue/" + book.h3.a['href']

            data_list.append({
                "ID": item_id,
                "Title": title,
                "Price (£)": float(price_text),
                "Rating (Out of 5)": rating,
                "Availability": availability,
                "URL": link
            })
            item_id += 1

    # 1. Export CSV
    with open("products.csv", mode="w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["ID", "Title", "Price (£)", "Rating (Out of 5)", "Availability", "URL"])
        writer.writeheader()
        writer.writerows(data_list)

    # 2. Export JSON
    with open("products.json", mode="w", encoding="utf-8") as f:
        json.dump(data_list, f, indent=4, ensure_ascii=False)

    # 3. Export Styled Excel (.xlsx)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "E-Commerce Products"

    headers = ["ID", "Title", "Price (£)", "Rating (Out of 5)", "Availability", "URL"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 26

    for item in data_list:
        ws.append([item["ID"], item["Title"], item["Price (£)"], item["Rating (Out of 5)"], item["Availability"], item["URL"]])

    data_font = Font(name="Segoe UI", size=10)
    wrap_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=len(data_list)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = data_font
            if cell.column in [1, 3, 4, 5]:
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 45

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    wb.save("products.xlsx")

    print(f"Success: Scraped {len(data_list)} products across {pages_to_scrape} pages!")
    print("Exported files: products.csv, products.json, products.xlsx")

if __name__ == "__main__":
    scrape_books(pages_to_scrape=3)